from datetime import datetime
from pathlib import Path

from flask import Flask, redirect, request, send_from_directory, url_for
from openpyxl import Workbook, load_workbook

app = Flask(__name__, static_folder='.', template_folder='.')
BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / 'messages.xlsx'


def save_message(name: str, email: str, subject: str, message: str) -> None:
    if EXCEL_FILE.exists():
        workbook = load_workbook(EXCEL_FILE)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Timestamp', 'Name', 'Email', 'Subject', 'Message'])

    worksheet.append([
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        name,
        email,
        subject,
        message,
    ])
    workbook.save(EXCEL_FILE)


@app.route('/')
def index():
    return send_from_directory(str(BASE_DIR), 'Muhammad_Hassaan_Zeb_Portfolio_v2.html')


@app.route('/Photos/<path:filename>')
def photos(filename):
    return send_from_directory(str(BASE_DIR / 'Photos'), filename)


@app.route('/submit-contact', methods=['POST'])
def submit_contact():
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    subject = request.form.get('subject', '').strip()
    message = request.form.get('message', '').strip()

    if not name or not email or not message:
        return 'Please provide your name, email, and a message.', 400

    save_message(name, email, subject, message)
    return redirect(url_for('thank_you', name=name))


@app.route('/thank-you')
def thank_you():
    name = request.args.get('name', 'Friend')
    return (
        '<!DOCTYPE html>'
        '<html lang="en">'
        '<head>'
        '  <meta charset="UTF-8">'
        '  <meta name="viewport" content="width=device-width, initial-scale=1.0">'
        '  <title>Thank You</title>'
        '  <style>'
        '    body{margin:0;font-family:Arial,sans-serif;background:#111315;color:#F5F7FA;display:flex;align-items:center;justify-content:center;height:100vh;text-align:center;padding:1rem;}'
        '    .card{max-width:520px;background:#1A1D21;border:1px solid #30363D;border-radius:18px;padding:2.5rem;box-shadow:0 24px 60px rgba(0,0,0,0.4);}'
        '    a{display:inline-block;margin-top:1.5rem;color:#39FF88;text-decoration:none;border:1px solid rgba(57,255,136,0.35);padding:0.7rem 1.25rem;border-radius:10px;transition:all 0.25s ease;}'
        '    a:hover{background:rgba(57,255,136,0.12);}'
        '  </style>'
        '</head>'
        '<body>'
        '  <div class="card">'
        f'    <h1>Thanks, {name}!</h1>'
        '    <p>Your message has been received and saved successfully.</p>'
        '    <a href="/">Return to portfolio</a>'
        '  </div>'
        '</body>'
        '</html>'
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
